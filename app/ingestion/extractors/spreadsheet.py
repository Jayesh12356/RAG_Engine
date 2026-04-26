"""Spreadsheet extractor — handles ``.xlsx``, ``.xls`` and ``.csv``.

For each sheet we emit two complementary views so retrieval works for both
"show me the table" (markdown-table window) and "what was APAC revenue"
(per-row natural-language sentence):

  * one ``kind="table"`` page per ~30-row window with the header repeated
  * one ``kind="row"`` page per row, capped at ``ROW_VIEW_CAP`` rows per file
    total to bound storage on huge exports
"""
from __future__ import annotations

import os

import structlog

from app.ingestion.pdf_parser import ParsedPage

logger = structlog.get_logger(__name__)

ROWS_PER_TABLE_WINDOW = 30
ROW_VIEW_CAP = 500
MAX_CELLS_PER_ROW = 40
MAX_CELL_CHARS = 200


def _safe_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if len(s) > MAX_CELL_CHARS:
        s = s[: MAX_CELL_CHARS - 1] + "…"
    return s.replace("\n", " ").replace("|", "\\|")


def _render_window(headers: list[str], rows: list[list[str]]) -> str:
    if not headers:
        return ""
    head_cells = [_safe_str(h) or f"col{i+1}" for i, h in enumerate(headers)]
    out: list[str] = []
    out.append("| " + " | ".join(head_cells) + " |")
    out.append("| " + " | ".join(["---"] * len(head_cells)) + " |")
    for r in rows:
        cells = [_safe_str(c) for c in r[: len(head_cells)]]
        cells = cells + [""] * max(0, len(head_cells) - len(cells))
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


def _row_to_sentence(sheet_name: str, headers: list[str], row: list, row_idx: int) -> str:
    parts: list[str] = []
    for h, v in zip(headers, row, strict=False):
        h_s = _safe_str(h)
        v_s = _safe_str(v)
        if h_s and v_s:
            parts.append(f"{h_s}: {v_s}")
    if not parts:
        return ""
    body = "; ".join(parts[:MAX_CELLS_PER_ROW])
    return f"Sheet '{sheet_name}' row {row_idx}: {body}."


def _iter_xlsx_sheets(path: str):
    """Yield ``(sheet_name, headers, rows)`` for each sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for .xlsx ingestion") from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                first = next(rows_iter)
            except StopIteration:
                continue
            headers = [_safe_str(c) for c in (first or [])]
            if not any(headers):
                continue
            data_rows: list[list] = []
            for raw in rows_iter:
                if raw is None:
                    continue
                row = list(raw)
                if not any((c is not None and str(c).strip()) for c in row):
                    continue
                data_rows.append(row)
            yield sheet_name, headers, data_rows
    finally:
        wb.close()


def _iter_xls_sheets(path: str):
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("xlrd is required for legacy .xls ingestion") from exc

    book = xlrd.open_workbook(path)
    for sheet in book.sheets():
        if sheet.nrows == 0:
            continue
        headers = [_safe_str(c) for c in sheet.row_values(0)]
        data_rows: list[list] = []
        for r in range(1, sheet.nrows):
            row = sheet.row_values(r)
            if not any((c is not None and str(c).strip()) for c in row):
                continue
            data_rows.append(list(row))
        yield sheet.name, headers, data_rows


def _iter_csv_sheets(path: str):
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("pandas is required for .csv ingestion") from exc

    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
    last_err: Exception | None = None
    df = None
    for enc in encodings:
        try:
            df = pd.read_csv(path, encoding=enc, dtype=object, keep_default_na=False, na_filter=False)
            break
        except Exception as exc:
            last_err = exc
            continue
    if df is None:
        raise RuntimeError(f"Failed to read CSV: {last_err}")
    headers = [_safe_str(c) for c in df.columns.tolist()]
    rows = df.values.tolist()
    yield os.path.splitext(os.path.basename(path))[0], headers, rows


def parse_spreadsheet(path: str) -> list[ParsedPage]:
    pdf_name = os.path.basename(path)
    service_name = os.path.splitext(pdf_name)[0].replace("_", " ")
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        sheet_iter = _iter_csv_sheets(path)
    elif ext == ".xls":
        sheet_iter = _iter_xls_sheets(path)
    else:
        sheet_iter = _iter_xlsx_sheets(path)

    pages: list[ParsedPage] = []
    page_no = 1
    rows_emitted_total = 0

    sheets = list(sheet_iter)
    total_sheets = max(1, len(sheets))

    for sheet_name, headers, rows in sheets:
        if not headers:
            continue
        # Markdown-table windows
        for start in range(0, max(1, len(rows)), ROWS_PER_TABLE_WINDOW):
            window = rows[start : start + ROWS_PER_TABLE_WINDOW]
            rendered = _render_window(headers, window)
            if not rendered:
                continue
            label = (
                f"{sheet_name} (rows {start + 1}-{start + len(window)})"
                if rows
                else sheet_name
            )
            text_with_header = f"# Sheet: {sheet_name}\n\n{rendered}"
            pages.append(
                ParsedPage(
                    page_number=page_no,
                    text=text_with_header,
                    pdf_name=pdf_name,
                    service_name=service_name,
                    section_title=label,
                    total_pages=total_sheets,
                    kind="table",
                )
            )
            page_no += 1

        # Per-row NL sentences (capped across the whole file)
        if rows_emitted_total < ROW_VIEW_CAP:
            remaining = ROW_VIEW_CAP - rows_emitted_total
            for i, row in enumerate(rows[:remaining], start=1):
                sentence = _row_to_sentence(sheet_name, headers, row, i)
                if not sentence:
                    continue
                pages.append(
                    ParsedPage(
                        page_number=page_no,
                        text=sentence,
                        pdf_name=pdf_name,
                        service_name=service_name,
                        section_title=f"{sheet_name} row {i}",
                        total_pages=total_sheets,
                        kind="row",
                    )
                )
                page_no += 1
                rows_emitted_total += 1
                if rows_emitted_total >= ROW_VIEW_CAP:
                    break

    if not pages:
        pages.append(
            ParsedPage(
                page_number=1,
                text="(empty spreadsheet)",
                pdf_name=pdf_name,
                service_name=service_name,
                section_title="Spreadsheet",
                total_pages=1,
                kind="prose",
            )
        )

    logger.info(
        "spreadsheet.parse.complete",
        path=path,
        sheets=total_sheets,
        pages=len(pages),
        rows_emitted=rows_emitted_total,
    )
    return pages
